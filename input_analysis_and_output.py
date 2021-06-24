# -*- coding: utf-8 -*-
"""
Created on Wed Jun 23 12:47:46 2021

@author: Administrator
"""


import openpyxl
import pandas as pd
import re
import string
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics.pairwise import cosine_similarity
import csv
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import easygui
import os

def main():
    URLinput = easygui.enterbox("Enter the Goodreads URL of the book for which you want recommendations: ")
    keyword_file = 'keyworded_concatenated.csv'
    df = pd.read_csv(keyword_file)
    indices = pd.Series(df['URL'])
    count = CountVectorizer()
    count_matrix = count.fit_transform(df['keywords'])
    cosine_sim = cosine_similarity(count_matrix, count_matrix)
    print(cosine_sim)
    recommended_books = []
    idx = indices[indices == URLinput].index[0]
    score_series = pd.Series(cosine_sim[idx]).sort_values(ascending = False)
    top_10_indices = list(score_series.iloc[1:11].index)
    top_10_index_scores = list(score_series.iloc[1:11])
    
    for i in top_10_indices:
        recommended_books.append(list(df['URL'])[i])
        
    
    df_metadata = pd.read_csv('./data/metadata.csv')
    
    df_new = pd.merge(df, df_metadata,
                      on = 'URL',
                      how = 'left')
        
    wb = Workbook()
    dest_filename = 'output.xlsx'
    ws1 = wb.active
    ws1.title = "Book Recommendations"
    
    ws1['A1']= 'Reference Book'
    ws1['A2'] ='URL'
    ws1['B2'] = 'Title'
    ws1['C2']= 'Author(s)'
    ws1['D2'] = 'Score'
    ws1['E2'] = 'Keywords Generated From 3-5 Star Reviews'
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 45
    ws1.row_dimensions[3].height = 35
    
    ws1['A3'].alignment = Alignment(horizontal='left',vertical='center',wrapText=True)
    ws1['B3'] = df_new[df_new.URL==URLinput].name.item()
    ws1['B3'].alignment = Alignment(horizontal='center',vertical='center')
    ws1['C3'] = df_new[df_new.URL==URLinput].authors.item()
    ws1['C3'].alignment = Alignment(horizontal='center',vertical='center')
    ws1['A3'] = URLinput
    ws1['D3'] = '1'
    ws1['D3'].alignment = Alignment(horizontal='center',vertical='center')
    ws1['E3'] = 'filler, text, keyword, descriptions, will, go, here'
    ws1['E3'].alignment = Alignment(horizontal='left',vertical='center',
                                    wrapText=True)
    ws1['A5'] = 'Recommendations'
    ws1['A6'] ='URL'
    ws1['B6'] = 'Title'
    ws1['C6']= 'Author(s)'
    ws1['D6'] = 'Score'
    ws1['E6'] = 'Keywords Generated From 3-5 Star Reviews'
    
    
    al_header = Alignment(horizontal='center',vertical='center')
    font_header = Font(bold=True)
    for row in ws1['A1:E2']:
        for cell in row:
            cell.font = font_header
            cell.alignment = al_header
    for row in ws1['A5:E6']:
        for cell in row:
            cell.font = font_header
            cell.alignment = al_header    
    
    
    
    ws1.merge_cells('A1:E1') 
    ws1.merge_cells('A5:E5') 

    j = 0
    for row in range(7, 17):
        cell = ws1.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal='left',vertical='center',wrapText=True)
        
        URL_out = recommended_books[j]
        score = top_10_index_scores[j]
        cell.value = URL_out
        
        name_out = df_new[df_new.URL==URL_out].name.item()
        cell2 = ws1.cell(row=row,column=2)
        cell2.alignment = Alignment(horizontal='center',vertical='center',
                                    wrapText=True)
        cell2.value = name_out
        
        author_out = df_new[df_new.URL==URL_out].authors.item()
        cell3 = ws1.cell(row=row,column=3)
        cell3.value = author_out
        cell3.alignment = Alignment(horizontal='center',vertical='center',
                                    wrapText=True)
        
        cell4 = ws1.cell(row=row,column=4)
        cell4.value = format(score,".3f")
        cell4.alignment = Alignment(horizontal='center',vertical='center')
        
        keywords_out = ""
        cell5 = ws1.cell(row=row,column=5)
        cell5.value = keywords_out
        cell5.alignment = Alignment(horizontal='left',vertical='center',
                                    wrapText=True)
        
        ws1.row_dimensions[row].height = 35
        
        j = j + 1
    
    wb.save(filename = dest_filename)
    
    os.startfile(dest_filename)
    
if __name__ == "__main__":
    main()